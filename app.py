import os
import logging
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy.exc import IntegrityError
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# Configure logging for debugging
logging.basicConfig(level=logging.DEBUG)

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")

# Configure the database
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Initialize the app with the extension
db.init_app(app)

# Import models after app initialization
from models import Product, ConsumptionBill, BillItem, ReceptionSheet, ReceptionItem, DraftBill, DraftBillItem, DraftReception, DraftReceptionItem

def init_db():
    """Initialize database tables"""
    with app.app_context():
        # Create all tables
        db.create_all()

@app.route('/')
def index():
    """Dashboard with low stock alerts and recent activity"""
    # Get low stock products
    low_stock_products = Product.query.filter(Product.quantity <= Product.min_stock).order_by(Product.quantity.asc()).all()
    
    # Get total products count
    total_products = Product.query.count()
    
    # Get recent consumption bills
    recent_bills = ConsumptionBill.query.order_by(ConsumptionBill.bill_date.desc()).limit(5).all()
    
    # Get recent receptions
    recent_receptions = ReceptionSheet.query.order_by(ReceptionSheet.reception_date.desc()).limit(5).all()
    
    return render_template('index.html', 
                         low_stock_products=low_stock_products,
                         total_products=total_products,
                         recent_bills=recent_bills,
                         recent_receptions=recent_receptions)

@app.route('/products')
def products():
    """Display all products with search functionality"""
    search_query = request.args.get('search', '')
    
    if search_query:
        products = Product.query.filter(
            (Product.code.ilike(f'%{search_query}%')) |
            (Product.name.ilike(f'%{search_query}%')) |
            (Product.location.ilike(f'%{search_query}%'))
        ).order_by(Product.name).all()
    else:
        products = Product.query.order_by(Product.name).all()
    
    return render_template('products.html', products=products, search_query=search_query)

@app.route('/products/add', methods=['GET', 'POST'])
def add_product():
    """Add new product"""
    if request.method == 'POST':
        code = request.form['code'].strip()
        name = request.form['name'].strip()
        unit = request.form['unit'].strip()
        quantity = float(request.form['quantity'])
        location = request.form['location'].strip()
        min_stock = float(request.form['min_stock'])
        
        try:
            product = Product(
                code=code,
                name=name,
                unit=unit,
                quantity=quantity,
                location=location,
                min_stock=min_stock
            )
            db.session.add(product)
            db.session.commit()
            flash('Produsul a fost adăugat cu succes!', 'success')
            return redirect(url_for('products'))
        except IntegrityError:
            db.session.rollback()
            flash('Codul produsului există deja!', 'error')
    
    return render_template('products.html', action='add')

@app.route('/products/edit/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    """Edit existing product"""
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        code = request.form['code'].strip()
        name = request.form['name'].strip()
        unit = request.form['unit'].strip()
        quantity = float(request.form['quantity'])
        location = request.form['location'].strip()
        min_stock = float(request.form['min_stock'])
        
        try:
            product.code = code
            product.name = name
            product.unit = unit
            product.quantity = quantity
            product.location = location
            product.min_stock = min_stock
            product.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Produsul a fost actualizat cu succes!', 'success')
            return redirect(url_for('products'))
        except IntegrityError:
            db.session.rollback()
            flash('Codul produsului există deja!', 'error')
    
    return render_template('products.html', action='edit', product=product)

@app.route('/products/delete/<int:product_id>')
def delete_product(product_id):
    """Delete product"""
    product = Product.query.get_or_404(product_id)
    db.session.delete(product)
    db.session.commit()
    
    flash('Produsul a fost șters cu succes!', 'success')
    return redirect(url_for('products'))

@app.route('/consumption_bills')
def consumption_bills():
    """Display all consumption bills"""
    bills = ConsumptionBill.query.order_by(ConsumptionBill.bill_date.desc()).all()
    return render_template('consumption_bills.html', bills=bills)

@app.route('/consumption_bills/create')
def create_consumption_bill():
    """Create new consumption bill"""
    # Load draft if exists
    draft_data = load_draft_bill()
    
    products = Product.query.order_by(Product.name).all()
    
    return render_template('bill_create.html', products=products, draft_data=draft_data)

@app.route('/consumption_bills/add_item', methods=['POST'])
def add_bill_item():
    """Add item to current bill (AJAX endpoint)"""
    product_code = request.form['product_code']
    quantity = float(request.form['quantity'])
    
    product = Product.query.filter_by(code=product_code).first()
    
    if not product:
        return jsonify({'error': 'Produsul nu a fost găsit'}), 400
    
    if quantity > product.quantity:
        return jsonify({'error': 'Cantitatea solicitată depășește stocul disponibil'}), 400
    
    # Add to session
    if 'bill_items' not in session:
        session['bill_items'] = []
    
    item_number = len(session['bill_items']) + 1
    item = {
        'item_number': item_number,
        'code': product.code,
        'name': product.name,
        'unit': product.unit,
        'quantity': quantity,
        'location': product.location
    }
    
    session['bill_items'].append(item)
    session.modified = True
    
    return jsonify({'success': True, 'item': item})

@app.route('/consumption_bills/remove_item/<int:item_index>')
def remove_bill_item(item_index):
    """Remove item from current bill"""
    if 'bill_items' in session and 0 <= item_index < len(session['bill_items']):
        session['bill_items'].pop(item_index)
        # Renumber items
        for i, item in enumerate(session['bill_items']):
            item['item_number'] = i + 1
        session.modified = True
        flash('Articolul a fost eliminat!', 'success')
    
    return redirect(url_for('create_consumption_bill'))

@app.route('/consumption_bills/save_draft', methods=['POST'])
def save_bill_draft():
    """Save current bill as draft"""
    employee_name = request.form.get('employee_name', '')
    employee_signature = request.form.get('employee_signature', '')
    
    # Clear existing draft
    DraftBill.query.delete()
    DraftBillItem.query.delete()
    
    # Save new draft
    draft = DraftBill(
        employee_name=employee_name,
        employee_signature=employee_signature
    )
    db.session.add(draft)
    db.session.flush()  # Get the ID
    
    # Save draft items
    if 'bill_items' in session:
        for item in session['bill_items']:
            draft_item = DraftBillItem(
                draft_id=draft.id,
                item_number=item['item_number'],
                product_code=item['code'],
                product_name=item['name'],
                unit=item['unit'],
                quantity=item['quantity'],
                location=item['location']
            )
            db.session.add(draft_item)
    
    db.session.commit()
    
    flash('Bonul a fost salvat ca ciornă!', 'success')
    return redirect(url_for('create_consumption_bill'))

@app.route('/consumption_bills/finalize', methods=['POST'])
def finalize_consumption_bill():
    """Finalize consumption bill"""
    employee_name = request.form['employee_name'].strip()
    employee_signature = request.form['employee_signature'].strip()
    
    if not employee_name:
        flash('Numele angajatului este obligatoriu!', 'error')
        return redirect(url_for('create_consumption_bill'))
    
    if 'bill_items' not in session or not session['bill_items']:
        flash('Nu există articole în bon!', 'error')
        return redirect(url_for('create_consumption_bill'))
    
    try:
        # Create bill
        bill = ConsumptionBill(
            employee_name=employee_name,
            employee_signature=employee_signature,
            is_finished=True
        )
        db.session.add(bill)
        db.session.flush()  # Get the ID
        
        # Add bill items and update stock
        for item in session['bill_items']:
            # Add item to bill
            bill_item = BillItem(
                bill_id=bill.id,
                item_number=item['item_number'],
                product_code=item['code'],
                product_name=item['name'],
                unit=item['unit'],
                quantity=item['quantity'],
                location=item['location']
            )
            db.session.add(bill_item)
            
            # Update product stock
            product = Product.query.filter_by(code=item['code']).first()
            if product:
                product.quantity -= item['quantity']
                product.updated_at = datetime.utcnow()
        
        # Clear draft
        DraftBill.query.delete()
        DraftBillItem.query.delete()
        
        db.session.commit()
        
        # Clear session
        session.pop('bill_items', None)
        
        flash('Bonul de consum a fost finalizat cu succes!', 'success')
        return redirect(url_for('consumption_bills'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Eroare la finalizarea bonului: {str(e)}', 'error')
        return redirect(url_for('create_consumption_bill'))

@app.route('/consumption_bills/view/<int:bill_id>')
def view_consumption_bill(bill_id):
    """View consumption bill details"""
    bill = ConsumptionBill.query.get_or_404(bill_id)
    items = BillItem.query.filter_by(bill_id=bill_id).order_by(BillItem.item_number).all()
    
    return render_template('bill_create.html', bill=bill, items=items, view_mode=True)

@app.route('/consumption_bills/export/<int:bill_id>')
def export_consumption_bill(bill_id):
    """Export consumption bill to Excel"""
    bill = ConsumptionBill.query.get_or_404(bill_id)
    items = BillItem.query.filter_by(bill_id=bill_id).order_by(BillItem.item_number).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bon Consum {bill_id}"
    
    # Header
    ws['A1'] = 'BON DE CONSUM'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Bill info
    ws['A3'] = f"Data: {bill.bill_date.strftime('%Y-%m-%d %H:%M')}"
    ws['A4'] = f"Angajat: {bill.employee_name}"
    ws['A5'] = f"Semnătura: {bill.employee_signature or '-'}"
    
    # Table header
    headers = ['Nr.', 'Cod Produs', 'Denumire', 'U.M.', 'Cantitate', 'Locație']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
    
    # Table data
    for row, item in enumerate(items, 8):
        ws.cell(row=row, column=1, value=item.item_number)
        ws.cell(row=row, column=2, value=item.product_code)
        ws.cell(row=row, column=3, value=item.product_name)
        ws.cell(row=row, column=4, value=item.unit)
        ws.cell(row=row, column=5, value=item.quantity)
        ws.cell(row=row, column=6, value=item.location or '-')
        
        # Add borders
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
    
    # Auto-adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'bon_consum_{bill_id}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/reception')
def reception():
    """Display all reception sheets"""
    receptions = ReceptionSheet.query.order_by(ReceptionSheet.reception_date.desc()).all()
    return render_template('reception.html', receptions=receptions)

@app.route('/reception/view/<int:reception_id>')
def view_reception(reception_id):
    """View reception details"""
    reception = ReceptionSheet.query.get_or_404(reception_id)
    items = ReceptionItem.query.filter_by(reception_id=reception_id).order_by(ReceptionItem.item_number).all()
    
    # Return JSON for AJAX request
    return jsonify({
        'reception': {
            'id': reception.id,
            'date': reception.reception_date.isoformat(),
            'supplier': reception.supplier,
            'document_number': reception.document_number,
            'notes': reception.notes,
            'is_finished': reception.is_finished
        },
        'items': [{
            'item_number': item.item_number,
            'product_code': item.product_code,
            'product_name': item.product_name,
            'unit': item.unit,
            'quantity': item.quantity,
            'location': item.location,
            'entry_date': item.entry_date.isoformat()
        } for item in items]
    })

@app.route('/reception/export/<int:reception_id>')
def export_reception(reception_id):
    """Export reception to Excel"""
    reception = ReceptionSheet.query.get_or_404(reception_id)
    items = ReceptionItem.query.filter_by(reception_id=reception_id).order_by(ReceptionItem.item_number).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Receptie {reception_id}"
    
    # Header
    ws['A1'] = 'FIȘĂ DE RECEPȚIE'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Reception info
    ws['A3'] = f"Data: {reception.reception_date.strftime('%Y-%m-%d %H:%M')}"
    ws['A4'] = f"Furnizor: {reception.supplier}"
    ws['A5'] = f"Nr. Document: {reception.document_number or '-'}"
    ws['A6'] = f"Observații: {reception.notes or '-'}"
    
    # Table header
    headers = ['Nr.', 'Cod Produs', 'Denumire', 'U.M.', 'Cantitate', 'Locație', 'Data Intrare']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
    
    # Table data
    for row, item in enumerate(items, 9):
        ws.cell(row=row, column=1, value=item.item_number)
        ws.cell(row=row, column=2, value=item.product_code)
        ws.cell(row=row, column=3, value=item.product_name)
        ws.cell(row=row, column=4, value=item.unit)
        ws.cell(row=row, column=5, value=item.quantity)
        ws.cell(row=row, column=6, value=item.location or '-')
        ws.cell(row=row, column=7, value=item.entry_date.strftime('%Y-%m-%d'))
        
        # Add borders
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
    
    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'receptie_{reception_id}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/reception/create')
def create_reception():
    """Create new reception sheet"""
    # Load draft if exists
    draft_data = load_draft_reception()
    
    products = Product.query.order_by(Product.name).all()
    
    return render_template('reception_create.html', products=products, draft_data=draft_data)

@app.route('/reception/add_item', methods=['POST'])
def add_reception_item():
    """Add item to current reception"""
    product_code = request.form['product_code']
    quantity = float(request.form['quantity'])
    
    product = Product.query.filter_by(code=product_code).first()
    
    if not product:
        return jsonify({'error': 'Produsul nu a fost găsit'}), 400
    
    # Add to session
    if 'reception_items' not in session:
        session['reception_items'] = []
    
    item_number = len(session['reception_items']) + 1
    item = {
        'item_number': item_number,
        'code': product.code,
        'name': product.name,
        'unit': product.unit,
        'quantity': quantity,
        'location': product.location
    }
    
    session['reception_items'].append(item)
    session.modified = True
    
    return jsonify({'success': True, 'item': item})

@app.route('/reception/remove_item/<int:item_index>')
def remove_reception_item(item_index):
    """Remove item from current reception"""
    if 'reception_items' in session and 0 <= item_index < len(session['reception_items']):
        session['reception_items'].pop(item_index)
        # Renumber items
        for i, item in enumerate(session['reception_items']):
            item['item_number'] = i + 1
        session.modified = True
        flash('Articolul a fost eliminat!', 'success')
    
    return redirect(url_for('create_reception'))

@app.route('/reception/save_draft', methods=['POST'])
def save_reception_draft():
    """Save current reception as draft"""
    supplier = request.form.get('supplier', '')
    document_number = request.form.get('document_number', '')
    notes = request.form.get('notes', '')
    
    # Clear existing draft
    DraftReception.query.delete()
    DraftReceptionItem.query.delete()
    
    # Save new draft
    draft = DraftReception(
        supplier=supplier,
        document_number=document_number,
        notes=notes
    )
    db.session.add(draft)
    db.session.flush()  # Get the ID
    
    # Save draft items
    if 'reception_items' in session:
        for item in session['reception_items']:
            draft_item = DraftReceptionItem(
                draft_id=draft.id,
                item_number=item['item_number'],
                product_code=item['code'],
                product_name=item['name'],
                unit=item['unit'],
                quantity=item['quantity'],
                location=item['location']
            )
            db.session.add(draft_item)
    
    db.session.commit()
    
    flash('Recepția a fost salvată ca ciornă!', 'success')
    return redirect(url_for('create_reception'))

@app.route('/reception/finalize', methods=['POST'])
def finalize_reception():
    """Finalize reception sheet"""
    supplier = request.form['supplier'].strip()
    document_number = request.form['document_number'].strip()
    notes = request.form['notes'].strip()
    
    if not supplier:
        flash('Furnizorul este obligatoriu!', 'error')
        return redirect(url_for('create_reception'))
    
    if 'reception_items' not in session or not session['reception_items']:
        flash('Nu există articole în recepție!', 'error')
        return redirect(url_for('create_reception'))
    
    try:
        # Create reception
        reception = ReceptionSheet(
            supplier=supplier,
            document_number=document_number,
            notes=notes,
            is_finished=True
        )
        db.session.add(reception)
        db.session.flush()  # Get the ID
        
        # Add reception items and update stock
        for item in session['reception_items']:
            # Add item to reception
            reception_item = ReceptionItem(
                reception_id=reception.id,
                item_number=item['item_number'],
                product_code=item['code'],
                product_name=item['name'],
                unit=item['unit'],
                quantity=item['quantity'],
                location=item['location']
            )
            db.session.add(reception_item)
            
            # Update product stock
            product = Product.query.filter_by(code=item['code']).first()
            if product:
                product.quantity += item['quantity']
                product.updated_at = datetime.utcnow()
        
        # Clear draft
        DraftReception.query.delete()
        DraftReceptionItem.query.delete()
        
        db.session.commit()
        
        # Clear session
        session.pop('reception_items', None)
        
        flash('Fișa de recepție a fost finalizată cu succes!', 'success')
        return redirect(url_for('reception'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Eroare la finalizarea recepției: {str(e)}', 'error')
        return redirect(url_for('create_reception'))

def load_draft_bill():
    """Load draft bill data"""
    draft = DraftBill.query.order_by(DraftBill.last_updated.desc()).first()
    
    if not draft:
        return None
    
    # Get draft items
    items = DraftBillItem.query.filter_by(draft_id=draft.id).all()
    
    # Convert to session format
    session['bill_items'] = []
    for item in items:
        session['bill_items'].append({
            'item_number': item.item_number,
            'code': item.product_code,
            'name': item.product_name,
            'unit': item.unit,
            'quantity': item.quantity,
            'location': item.location
        })
    
    return {
        'employee_name': draft.employee_name,
        'employee_signature': draft.employee_signature
    }

def load_draft_reception():
    """Load draft reception data"""
    draft = DraftReception.query.order_by(DraftReception.last_updated.desc()).first()
    
    if not draft:
        return None
    
    # Get draft items
    items = DraftReceptionItem.query.filter_by(draft_id=draft.id).all()
    
    # Convert to session format
    session['reception_items'] = []
    for item in items:
        session['reception_items'].append({
            'item_number': item.item_number,
            'code': item.product_code,
            'name': item.product_name,
            'unit': item.unit,
            'quantity': item.quantity,
            'location': item.location
        })
    
    return {
        'supplier': draft.supplier,
        'document_number': draft.document_number,
        'notes': draft.notes
    }

# Initialize database on startup
init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)